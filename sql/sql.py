import sqlite3 as sql
from sqlite3 import Error
import openpyxl
from openpyxl import Workbook

def create_connection(path):
    connection = None
    try:
        connection = sql.connect(path)
        print("Connection to SQLite DB successful")
    except Error as e:
        print(f"The error {e} is occurred")

    return connection

def execute_query(connection, query):
    cursor = connection.cursor()
    try:
        cursor.execute(query)
        connection.commit()
        print("Query executed successfully")
    except Error as e:
        print(f"The error '{e}' is occured")

# query - передаваемый запрос

create_users_table = '''
CREATE TABLE IF NOT EXISTS users (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    age INTEGER,
    gender TEXT,
    nationality TEXT
);
'''

create_posts_table = '''
CREATE TABLE IF NOT EXISTS posts (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    title TEXT NOT NULL,
    description TEXT NOT NULL,
    user_id INTEGER NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users (id)
);
'''

create_comments_table = '''
CREATE TABLE IF NOT EXISTS comments (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    text TEXT NOT NULL,
    user_id INTEGER NOT NULL,
    post_id INTEGER NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users (id) FOREIGN KEY (post_id) REFERENCES posts (id)
);
'''

create_likes_table = '''
CREATE TABLE IF NOT EXISTS likes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    post_id INTEGER NOT NULL,
    FOREIGN KEY (user_id) REFERENCES users (id) FOREIGN KEY (post_id) REFERENCES posts (id)
);
'''

connection = create_connection("sm_app.sqlite")
execute_query(connection, create_users_table)
execute_query(connection, create_posts_table)
execute_query(connection, create_comments_table)
execute_query(connection, create_likes_table)

create_users = '''
INSERT INTO
    users (name, age, gender, nationality)
VALUES
    ('Petrovich777', 20, 'male', 'Germany'),
    ('Egorka', 21, 'male', 'Italy'),
    ('Fedot_Danetot', 68, 'male', 'Russia'),
    ('ddaaasshhaaa', 18, 'female', 'Russia'),
    ('Falafel_812', 27, 'female', 'Russia');
'''

execute_query(connection, create_users)

create_posts = '''
INSERT INTO
    posts (title, description, user_id)
VALUES
    ("Happy", "I am feeling very happy today", 1),
    ("Hot Weather", "The weather is very hot today", 2),
    ("Help", "I need some help with my work", 2),
    ("Great News", "I am getting married", 1),
    ("Interesting Game", "It was a fantastic game of tennis", 5),
    ("Party", "Anyone up for a late-night party today?", 3);
'''

execute_query(connection, create_posts)

create_comments = """
INSERT INTO
  comments (text, user_id, post_id)
VALUES
  ('Count me in', 1, 6),
  ('What sort of help?', 5, 3),
  ('Congrats buddy', 2, 4),
  ('I was rooting for Nadal though', 4, 5),
  ('Help with your thesis?', 2, 3),
  ('Many congratulations', 5, 4);
"""

create_likes = """
INSERT INTO
  likes (user_id, post_id)
VALUES
  (1, 6),
  (2, 3),
  (1, 5),
  (5, 4),
  (2, 4),
  (4, 2),
  (3, 6);
"""

execute_query(connection, create_comments)
execute_query(connection, create_likes)

def execute_read_query(connection, query):
    cursor = connection.cursor()
    result = None
    try:
        cursor.execute(query)
        result = cursor.fetchall()
        return result
    except Error as e:
        print(f"The error '{e}' is occured")

select_users = "SELECT * from users"
users = execute_read_query(connection, select_users)

for user in users:
    print(user)

select_posts = "SELECT * FROM posts"
posts = execute_read_query(connection, select_posts)

for post in posts:
    print(post)

select_users_posts = '''
SELECT 
    users.id,
    users.name,
    posts.description
FROM
    posts
    INNER JOIN users ON users.id = posts.user_id
'''

users_posts = execute_read_query(connection, select_users_posts)

for users_post in users_posts:
    print(users_post)

select_posts_comments_users = '''
SELECT
    posts.description as post,
    text as comment,
    name
FROM
    posts
    INNER JOIN comments ON posts.id = comments.post_id
    INNER JOIN users ON users.id = comments.user_id
'''

posts_comments_users = execute_read_query(
    connection, select_posts_comments_users
)

for posts_comments_user in posts_comments_users:
    print(posts_comments_user)

cursor = connection.cursor()
cursor.execute(select_posts_comments_users)
cursor.fetchall()

column_names = [description[0] for description in cursor.description]
print(column_names)

select_post_likes = '''
SELECT
    description as Post,
    COUNT(likes.id) as Likes
FROM
    likes,
    posts
WHERE
    posts.id = likes.post_id
GROUP BY
    likes.post_id
'''

post_likes = execute_read_query(connection, select_post_likes)

for post_like in post_likes:
    print(post_like)

select_post_description = "SELECT description FROM posts WHERE id = 2"

post_description = execute_read_query(connection, select_post_description)

for description in post_description:
    print(description)

update_post_description = """
UPDATE
    posts
SET
    description = 'The weather has become pleasant now'
WHERE
    id = 2
"""

execute_query(connection, update_post_description)

delete_comment = "DELETE FROM comments WHERE id = 5"
execute_query(connection, delete_comment)

wb = Workbook()
wb.create_sheet("users")
wb.create_sheet("posts")
wb.create_sheet("comments")
wb.create_sheet("likes")

ws = wb["users"]
cursor.execute("SELECT * FROM users")
users_data = cursor.fetchall()

for row_index, row_data in enumerate(users_data, start=1):
    for col_index, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_index, column=col_index, value=cell_data)

ws = wb["posts"]
cursor.execute("SELECT * FROM posts")
users_data = cursor.fetchall()

for row_index, row_data in enumerate(users_data, start=1):
    for col_index, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_index, column=col_index, value=cell_data)

ws = wb["comments"]
cursor.execute("SELECT * FROM comments")
users_data = cursor.fetchall()

for row_index, row_data in enumerate(users_data, start=1):
    for col_index, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_index, column=col_index, value=cell_data)

ws = wb["likes"]
cursor.execute("SELECT * FROM likes")
users_data = cursor.fetchall()

for row_index, row_data in enumerate(users_data, start=1):
    for col_index, cell_data in enumerate(row_data, start=1):
        ws.cell(row=row_index, column=col_index, value=cell_data)

wb.save('sm_app.xlsx')

connection.close()